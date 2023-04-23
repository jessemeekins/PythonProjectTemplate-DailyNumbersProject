"""
Copyright (c) 2023 Jesse Meekins
See project 'license' file for more informations
"""

from abc import ABC, abstractmethod
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook
import string

class ReportFactory(ABC):
    def __init__(self, data, *args, **kwargs):
        self.data = data

    # MAIN EXECUTION
    def generate_report(self, filename):
        workbook = Workbook()
        sheet = workbook.active

        self.create_report(sheet)
        workbook.save(filename)
    
    @abstractmethod
    def create_report(self, sheet):
        pass

    # Class Variables 
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    light_red_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    header_font = Font(size=14, bold=True)
    subheader_font = Font(size=12, italic=True)


    def _add_header_and_subheader(self, sheet, headers, sub_headers):
        for cell, value in headers:
            sheet[cell] = value
            sheet[cell].font = self.header_font
            sheet[cell].fill = self.light_green_fill
            sheet[cell].alignment = Alignment(horizontal='center')

        for cell, value in sub_headers:
            sheet[cell] = value
            sheet[cell].font = self.subheader_font
    
    # CLASS FUNCTIONS
    def _add_report_title(self, sheet, title, cell="B1"):
        sheet.merge_cells('B1:P1')
        title_cell = sheet[cell]
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, italic=True)
        title_cell.fill = self.light_red_fill
        title_cell.alignment = Alignment(horizontal='center')

    def _populate_chiefs_data(self, sheet):
        chief_data = self.data.get('chiefs')
        # Styles
        header_font = self.header_font
        header_fill = self.light_green_fill

        divisions = {
            "DIVISION 1 CHIEFS": {
                "merger": "E4:F4",
                "header_col": "E",
                "data_col": "F",
                "chiefs": [
                    ("DC001", "BC001", "BC002", "BC003", "BC006", "BC008", "BC009"),
                ],
            },
            "DIVISION 2 CHIEFS": {
                "merger": "H4:I4",
                "header_col": "H",
                "data_col": "I",
                "chiefs": [
                    ("DC002", "BC004", "BC005", "BC007", "BC010", "BC011"),
                ],
            },
        }

        row_start = 4
        for division_name, division_info in divisions.items():
            sheet.merge_cells(f"{division_info['merger']}")
            header_cell = sheet[f'{division_info["header_col"]}{row_start}']
            header_cell.value = division_name
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal='center')

            for i, chief_id in enumerate(division_info["chiefs"][0], start=row_start + 1):
                sheet[f'{division_info["header_col"]}{i}'] = chief_id
                sheet[f'{division_info["header_col"]}{i}'].font = Font(bold=True)
                #if i == row_start + 1:
                sheet[f'{division_info["data_col"]}{i}'] = chief_data.get(chief_id)

        # Auto size column widths
        for col_letter in ['E', 'F', 'H', 'I']:
            sheet.column_dimensions[col_letter].auto_size = True

    
    def _add_als_company_data(self, sheet):
        als_data = self.data.get('als')
        _length_of_off = self.data.get("off_by_rank")
        l = len(_length_of_off) + 4
        sheet.merge_cells(f'H{l+3}:P{l+3}')
        sheet.merge_cells(f'J{l+4}:K{l+4}')

        # Add the header row
        headers = [
            (f"H{l+3}", "ALS COMPANIES"),
        ]
        
        sub_headers = [
            (f"H{l+4}", "BATTLAION"),
            (f"I{l+4}", "COUNT"),
            (f"J{l+4}", "TOTAL COUNT:"),
            (f"L{l+4}", als_data["count"]),
         ]

        battalions = [
            (f"H{l+5}", "BATT 1"),
            (f"H{l+6}", "BATT 2"),
            (f"H{l+7}", "BATT 3"),
            (f"H{l+8}", "BATT 6"),
            (f"H{l+9}", "BATT 8"),
            (f"H{l+10}", "AR-C"),
            (f"H{l+11}", "BATT 9"),
            (f"H{l+12}", "BATT 4"),
            (f"H{l+13}", "BATT 5"),
            (f"H{l+14}", "BATT 7"),
            (f"H{l+15}", "BATT 10"),
            (f"H{l+16}", "BATT 11"),
        ]

        self._add_header_and_subheader(sheet, headers, sub_headers)

        
        for cell, value in battalions:
            sheet[cell] = value
            sheet[cell].font = Font(bold=True)

        companies = als_data["companies"]
        num = l+5
        start_letter = 'J'
        start_index = ord(start_letter) - ord('A')

        for company in companies.values():
            new_comp = list(company)
            for i, letter in enumerate(string.ascii_uppercase[start_index:start_index+len(company)]):
                sheet[f"{letter}{num}"] = new_comp[i]
            num += 1


    def _add_paycode_data(self, sheet):
        paycode_data = self.data.get('paycodes')
        
        
        sheet.merge_cells('B4:C4')

        # Headers
        headers = [
            ("B4", "PAYCODES"),
        ]
        sub_headers = [
            ('B5', "CODE"),
            ('C5', "COUNT")
        ]

        self._add_header_and_subheader(sheet, headers, sub_headers)
            
        start_row = 6
        for i, (k, v) in enumerate(paycode_data.items(), start=start_row):
            sheet[f"B{i}"] = k
            sheet[f"C{i}"] = v

        # Auto size column widths
        sheet.column_dimensions["B"].auto_size = True
        sheet.column_dimensions["C"].auto_size = True

    def _off_sl_more_than_two_days(self, sheet):
        _length_of_off = self.data.get("off_by_rank")
        length_of_of_by_ranks = len(_length_of_off) + 4
        long_term = self.data.get("multi_day")
        sheet.merge_cells(F'E{length_of_of_by_ranks+3}:F{length_of_of_by_ranks+3}')

        # Headers
        headers = [
            (f"E{length_of_of_by_ranks+3}", "LONG TERM SICK LEAVE"),
        ]
        sub_headers = [
            (f"E{length_of_of_by_ranks+4}", "NAME"),
            (f"F{length_of_of_by_ranks+4}", "DAYS"),
        ]

        self._add_header_and_subheader(sheet, headers, sub_headers)

        start_row = length_of_of_by_ranks+5
        for i, (k, v) in enumerate(long_term.items(), start=start_row):
            sheet[f"E{i}"] = k
            sheet[f"F{i}"] = v["days_off"]

        # Auto size column widths
        sheet.column_dimensions["E"].auto_size = True
        sheet.column_dimensions["G"].auto_size = True
    
    def _add_off_by_ranks(self, sheet):
        off_by_rank = self.data.get('off_by_rank')

        sheet.merge_cells("K4:L4")

        headers = [
            ("K4", "OFF BY RANK")
        ]
        sub_headers = [
            ("K5", "RANK"),
            ("L5", "COUNT")
        ]

        self._add_header_and_subheader(sheet, headers, sub_headers)

        start_row = 6
        for i, (k, v) in enumerate(off_by_rank.items(), start=start_row):
            sheet[f"K{i}"] = k
            sheet[f"L{i}"] = v
    
class DailyNumbersReport(ReportFactory):
    def create_report(self, sheet):
        self._add_report_title(sheet, "Daily Numbers")
        self._populate_chiefs_data(sheet)
        self._add_als_company_data(sheet)
        self._add_paycode_data(sheet)
        self._off_sl_more_than_two_days(sheet)
        self._add_off_by_ranks(sheet)




