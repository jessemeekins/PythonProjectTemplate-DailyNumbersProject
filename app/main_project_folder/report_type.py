#%%
"""
Copyright (c) 2023 Jesse Meekins
See project 'license' file for more informations
"""
import json
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from typing import (
    Dict, List, Set, 
    Tuple, Union, Any
)

from main_project_folder.class_functions import ReportType

class CurrentShift(ReportType):
    def _get_shift(self):
        shift = self.get_current_shift(self.data)
        return shift
    
    def __call__(self):
        shift = self._get_shift()
        return {"shift": shift}

    
class AlsCompanyReport(ReportType):
    def _is_als_company(self, record: Dict[str, str]) -> bool:
        is_on_duty = self.on_duty(record)
        working_right_now = self.right_now(record)
        is_paramedic = self.is_paramedic(record)
        company = self.get_company(record)
        
        if is_on_duty and is_paramedic and working_right_now:
            if any(x in company for x in ['PU0', 'TR0', 'RC0', 'QT0']):
                return True
        return False
        
    def list_als_companies(self) -> Dict[str, Set[str]]:
        als_set = dict()
        als_json = dict()
        for record in self.data.values():
            company_is_als = self._is_als_company(record)
            if company_is_als:
                region = self.get_region(record)
                als_set.setdefault(region, set()).add(self.get_company(record))
                als_json[self.get_company(record)] = {"region": self.get_region(record), "is_als": True}
        return als_set, als_json

    def json_als_companies(self, companies: dict) -> dict:
        with open('/Users/jessemeekins/Documents/VS_CODE/PythonProjectTemplate/app/data/imports/ALS_JSON.json', 'w') as f:
            json.dump(companies, f)
            print('Done')
    

    @staticmethod
    def als_count(als_companies: Dict[str, Set[str]]) -> int:
        company_list = [iter for values in als_companies.values() for iter in values]
        return company_list

    def __call__(self) -> Dict[str, Union[Dict[str, Set[str]], int]]:
        company_dict, json_dict = self.list_als_companies()
        self.json_als_companies(json_dict)
        return {"als_companies":company_dict}


class AssignmentReport(ReportType):
    def __init__(self, data: Dict[str, Dict[str, Union[str, List[str]]]]) -> None:
        self.data: Dict[str, Dict[str, Union[str, List[str]]]] = data

    def all_records(self) -> List[Tuple[str, str, str, str, str, str, List[str]]]:
        records = []
        for record in self.data.values():
            location = record["location"]
            if location == "FIELD":
                shift = self._get_shift_specialty(record)
                company = self.get_specialty_company(record)
                for battalion, companies in ComplimentReport.BATTALION_DICT.items():
                    if company in companies["companies"]:
                        assigned_region = battalion
                        records.append((shift, assigned_region, company, record["eid"], record["rank"], record["last_name"], record["first_name"]))
        return records
    
    def group_assignments(self):
        personnel = self.all_records()
        assignments = {}
        
        for person in personnel:
            shift, region, company, eid, rank, last_name, first_name = person
            division = region[:3]
            battalion = region[-3:]
            
            # Create the nested dictionary structure
            if shift not in assignments:
                assignments[shift] = {}
            if division not in assignments[shift]:
                assignments[shift][division] = {}
            if battalion not in assignments[shift][division]:
                assignments[shift][division][battalion] = {}
            if company not in assignments[shift][division][battalion]:
                assignments[shift][division][battalion][company] = []
            
            # Add the person to the nested dictionary
            assignments[shift][division][battalion][company].append({"eid": eid, "rank": rank, "last": last_name, "first": first_name})
            
        return assignments

  
    def save_assignments_to_excel(self, assignments: dict, filename: str, grouped=False):
        # Create an empty DataFrame to hold the results
        df = pd.DataFrame(columns=["Shift", "Division", "Battalion", "Company", "EID", "Rank", "Last", "First"])

        # Loop through each level of the nested dictionary and add the data to the DataFrame
        for shift, shift_data in assignments.items():
            for div, div_data in shift_data.items():
                for batt, batt_data in div_data.items():
                    for comp, people in batt_data.items():
                        for person in people:
                            df = df.append({"Shift": shift, "Division": div, "Battalion": batt, "Company": comp,
                                            "EID": person["eid"], "Rank": person["rank"], "Last": person["last"], "First": person["first"]},
                                        ignore_index=True)

        # Save the DataFrame to Excel
        if grouped == False:
            writer = pd.ExcelWriter(filename, engine="openpyxl")
            df.to_excel(writer, index=False, sheet_name="Assignments")
            workbook = writer.book
            worksheet = writer.sheets["Assignments"]

            # Add styling
            header_font = Font(bold=True)
            alignment = Alignment(horizontal="center")

            for col_letter in range(ord("A"), ord("H") + 1):
                col_letter = chr(col_letter)
                header_cell = worksheet[col_letter + "1"]
                header_cell.font = header_font
                header_cell.alignment = alignment
                worksheet.column_dimensions[col_letter].width = 15

            # Save the Excel file
            writer.save()
            writer.close()
        
        if grouped == True:

            # Group the data by columns
            grouped = df.groupby(["Shift", "Division", "Battalion", "Company"])

            # Create a new Excel workbook and get the active sheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Assignments"

            # Add styling to the header row
            header_font = Font(bold=True)
            alignment = Alignment(horizontal="center")

            for col_num, column_name in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_name
                cell.font = header_font
                cell.alignment = alignment
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_num).column_letter].width = 15

            # Write the grouped data to separate worksheets
            row_num = 2  # Start writing from the second row
            for group_name, group_data in grouped:
                group_data.reset_index(drop=True, inplace=True)  # Reset the index for each group
                for col_num, column_value in enumerate(group_name, start=1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_value
                for col_num, column_value in enumerate(group_data.columns, start=len(group_name) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    try:
                        cell.value = group_data[column_value]
                    except:
                        pass
                row_num += 1

            # Save the Excel file
            workbook.save(filename)
                    
    
    def __call__(self) -> List[Tuple[str, str, str, str, str, str, List[str]]]:
        assignments = self.group_assignments()
        self.save_assignments_to_excel(assignments, '/Users/jessemeekins/Desktop/assignments.xlsx')
        #data = self.all_records()
        return {"data":assignments}


class OnDutyChiefs(ReportType):
    def __init__(self, data:dict) -> None:
        self.data = data 
    
    def on_duty_chiefs(self) -> Dict[str, str]:
        chiefs = dict()
        for value in self.data.values():
            is_on_duty = self.on_duty(value)
            is_a_chief = self.bc_or_dc_working(value)
            if is_on_duty and is_a_chief:
                rank = value["rank"]
                name = value["name"]
                company = value["company"]
                if rank == "BC-FIRE" or "BC-EMS" or "BC-ARC":
                    chiefs[company] = name
                else:
                    pass
        return chiefs
    
    def __call__(self):
        chiefs = self.on_duty_chiefs()
        return {"chief":chiefs}


class PayCodes(ReportType):

    PAYCODE_GROUPS = [
        ("Scheduled Day", r"SD|SDACT|WSB"),
        ("Vacation Leave", r"VL|VCALL|SL_VL"),
        ("Overtime", r"OT(CB|HO|OR)?|EXT"),
        ("Bereavment", r"BV"),
        ("Suspended", r"SWOP|SWP"),
        ("Sick Leave", r"SL"),
        ("LWOP", r"LWOP|FLWOP"),
        ("OJI", r"OJI"),
        ("LD", r"LD"),
    ]

    def __init__(self, data: Dict[str, Dict[str, str]]) -> None:
        self.data = data
        self.current_shift = self.get_current_shift(self.data)

    def is_supposed_to_be_at_work(self, record) -> bool:
        is_shift = self.get_shift_specialty(self.current_shift, record)
        is_field = self.is_field_compliment(record)
        right_now = self.right_now(record)
        is_working = self.on_duty(record)

        if is_field and is_shift and is_working and right_now:
            return True
        
    def get_available(self) -> dict:
        available_personnel = list()
        for record in self.data.values():
            is_true = self.is_supposed_to_be_at_work(record)
            if is_true:
                available_personnel.append(record)
        return len(available_personnel)
       
    def payroll_code_count(self) -> Dict[str, List[Dict[str, str]]]:
        grouped = dict()
        TODAYS_SHIFT = self.current_shift
        
        for record in self.data.values():
            is_field = self.region_in_field(record)
            right_now = self.right_now(record)
            paycode = record["paycode"]

            if is_field and right_now:
                for group in self.PAYCODE_GROUPS:
                    pattern = group[1]
                    if re.search(pattern, paycode):
                        if group[0] not in grouped.keys():
                            grouped[group[0]] = set()
                        grouped[group[0]].add(record["name"])

        return grouped

    def paycode_count(self) -> Dict[str, int]:
        count_dict = dict()
        grouped_paycodes = self.payroll_code_count()
        for paycode, lst in grouped_paycodes.items():
            count_dict[paycode] = len(lst)
        return count_dict
    
    def total_people(self) -> str:
        num_working = list()
        for value in self.data.values():
            on_duty = self.on_duty(value)
            right_now = self.right_now(value)
            is_field = self.region_in_field(value)
            has_shift = self.twenty_four_hour_shift(value)

            if on_duty and right_now and is_field and has_shift:
                num_working.append(value["shift"])
        return len(num_working)
                
    def __call__(self):
        available = self.get_available()
        num = self.total_people()
        paycode = self.paycode_count()
        print(num, available)
        return {"paycode": paycode}
    

class RankCounts(ReportType):
    def __init__(self, data: dict[str,str]) -> None:
        self.data = data

    def off_by_rank(self) -> Dict[str,Dict[str, str]]:
        off_by_rank = dict()
        for value in self.data.values():
            is_field = self.region_in_field(value)
            off_duty = self.off_duty(value)
            is_vcall = self.is_on_vcall(value)

            if off_duty or is_vcall:
                if is_field:
                    rank = value["rank"]
                    if rank[0] == '.':
                        rank = rank[1:]
                    else:
                        rank
                    if rank not in off_by_rank.keys():
                        off_by_rank[rank] = []
                    off_by_rank[rank].append(value["name"])
        return off_by_rank

    def off_by_rank_count(self) -> Dict[str, int]:
        off_count = dict()
        ranks = self.off_by_rank()
        for rank, lst in ranks.items():
            off_count[rank] = len(lst)
        return off_count
    
    def __call__(self) -> Dict[str, int]:
        off_by_rank_count = self.off_by_rank_count()
        return {"ranks": off_by_rank_count}


class ComplimentReport(ReportType):
    BATTALION_DICT = {
        "D01B01": {"companies":["PU001", "PU002", "PU005", "PU007", "PU008", "TR002", "TR005", "TR013", "BC001", "DC001"], "required":31, "compliment": dict(), "available": set(), "balance": int()},
        "D01B02": {"companies":["PU004", "PU015", "PU019", "PU026", "PU028", "TR011", "BC002"], "required":37, "compliment": dict(), "available": set(), "balance": int()},
        "D01B03": {"companies":["PU010", "PU014", "PU020", "PU029", "PU032", "TR009", "RC002", "BC003"], "required":36, "compliment": dict(), "available": set(), "balance": int()},
        "D01B06": {"companies":["PU011", "PU013", "PU016", "PU018", "PU022", "TR004", "TR007", "BC006", "AT001", "RH001"], "required":37, "compliment": dict(), "available": set(), "balance": int()},
        "D01B08": {"companies":["PU034", "PU040", "PU042", "PU050", "QT057", "TR021", "BC008"], "required":31, "compliment": dict(), "available": set(), "balance": int()},
        "D01B09": {"companies":["PU036", "QT037", "PU038", "PU039", "PU043", "PU045", "TR018", "TR019", "TR024", "BC009"], "required":41, "compliment": dict(), "available": set(), "balance": int()},
        "D01BAC": {"companies":["PU033", "TR016", "AR001", "AR002", "AR003", "AC001"], "required":20, "compliment": dict(), "available": set(), "balance": int()},
        "D02B04": {"companies":["QT054", "PU056", "PU058", "PU059", "TR030", "BC004"], "required":25, "compliment": dict(), "available": set(), "balance": int()},
        "D02B05": {"companies":["PU017", "PU023", "PU024", "QT048", "PU051", "TR008", "TR010", "BC005"], "required":37, "compliment": dict(), "available": set(), "balance": int()},
        "D02B07": {"companies":["PU021", "PU025", "PU030", "PU041", "PU044", "TR015", "TR020", "RC001", "BC007", "DC002"], "required":42, "compliment": dict(), "available": set(), "balance": int()},
        "D02B10": {"companies":["PU035", "PU052", "PU053", "PU055", "TR017", "TR027", "BC010"], "required":31, "compliment": dict(), "available": set(), "balance": int()},
        "D02B11": {"companies":["PU027", "PU031", "PU046", "PU047", "PU049", "TR023", "RC003", "BC011"], "required":33, "compliment": dict(), "available": set(), "balance": int()},
        "EMS": {"companies":["BC020", "ES201", "ES202", "ES203", "ES204", "ES205"], "required": 6, "compliment": dict(), "available": set(), "balance": int()}
    }

    def __init__(self, data: dict) -> None:
        self.data = data
        self.current_shift = self.get_current_shift(self.data)

    def get_all_field_personnel(self) -> None:
        for value in self.data.values():
            has_shift = self.get_shift_specialty(self.current_shift, value)
            is_field = self.is_field_compliment(value)
            

            if has_shift and is_field:
                company = self.get_specialty_company(value)
                for battalion, companies in self.BATTALION_DICT.items():
                    
                    if company in companies["companies"]:
                        
                        assigned_region = battalion
                        break

                else:
                    print(ValueError(f"Could not find battalion for {company}"))
                    ...
                self.BATTALION_DICT[assigned_region]["compliment"][value["name"]] = value

                
                
        
    def get_available_personnel(self):
        for value in self.data.values():
            has_shift = self.get_shift_specialty(self.current_shift, value)
            right_now = self.right_now(value)
            on_duty = self.on_duty(value)

            if has_shift and right_now and on_duty:
                comp = self.get_specialty_company(value)
                
            for battalion ,companies in self.BATTALION_DICT.items():
                if comp in companies:
                    region = battalion

            try:
                if region in self.BATTALION_REQUIRED_AND_COMPLIMENT.keys() and region in self.BATTALION_DICT.keys():
                    self.BATTALION_REQUIRED_AND_COMPLIMENT[region] = set()
                for battalion, companies in self.BATTALION_DICT.items():
                    if comp in companies["companies"]:
                        self.BATTALION_REQUIRED_AND_COMPLIMENT[battalion]["available"].add(value)
            except:
                ...
        


    def count_compliment(self):
        count_dict = dict()
        grouped = self.get_all_field_personell()
        for batt, lst in grouped.items():
            min_req = self.BATTALION_REQUIRED_AND_COMPLIMENT.get(batt, 0)
            count_dict[batt]["compliment"] = (len(lst), min_req)
        return count_dict

    def count_compliment_and_update_dict(self):
        for battalion, companies in self.BATTALION_DICT.items():
            compliment = companies["compliment"]
            available = companies["availabe"]

            self.BATTALION_DICT[battalion]["compliment"] = len(compliment)
    
    def count_available(self):
        available = dict()

    def __call__(self):
        self.get_all_field_personnel()

        return {"data":self.BATTALION_DICT}
    

class DetailedPersonnel(ReportType):
    def __init__(self, data) -> None:
        self.data = data
    def detailed_personnel(self):
        TODAYS_SHIFT = self.get_current_shift(self.data)

        detailed_personnel = dict()
        for record in self.data.values():
  
            is_field = self.is_field_compliment(record)
            is_right_now = self.right_now(record)
            if is_field and is_right_now:
                # Retrieve current roster data for each staffing record
                # Returns strings
                name: str = self.get_name(record)
                rank: str = self.get_rank(record)
                current_region: str = self.get_region(record)
                current_company: str = self.get_company(record)
                # Retrieve SPeciality information
                specialty_company: str = self.get_specialty_company(record)
                specialty_shift: str = self._get_shift_specialty(record)
                is_off = self.off_duty(record)
                assigned_region: str = None
                is_on_vcall = self.is_on_vcall(record)



                for battalion, companies in ComplimentReport.BATTALION_DICT.items():
                    if specialty_company in companies["companies"]:
                        assigned_region = battalion
                                 

                        if assigned_region != current_region and TODAYS_SHIFT == specialty_shift:
                            if assigned_region not in detailed_personnel:
                                detailed_personnel[assigned_region] = []
                            
                            detailed_personnel[assigned_region].append((rank, name, current_region, current_company, record["shift"], record['paycode']))
                            ComplimentReport.BATTALION_DICT[assigned_region]['available'].add(name)
                        
                        if TODAYS_SHIFT == specialty_shift:
                            if is_off or is_on_vcall:
                                if assigned_region not in detailed_personnel:
                                    detailed_personnel[assigned_region] = []
                                detailed_personnel[assigned_region].append((rank, name, current_region, current_company, record["shift"], record["paycode"]))
                                ComplimentReport.BATTALION_DICT[assigned_region]['available'].add(name)
                                ...

        return detailed_personnel
    
    def multiple_days_off(self) -> Dict[str, Dict[str, Union[str, str]]]:
        sl = dict()
        for value in self.data.values():
            is_field = self.region_in_field(value)
            greater_than_two = self.find_numbers_greater_than_two(value)

            if value["paycode"] == "SL":

                if value["detail_code"] != 'Not listed' and is_field and greater_than_two:
                    sl[value["name"]] = {"paycode": value["paycode"], "days_off": greater_than_two[0]}
        return sl

    def __call__(self):
        print(ComplimentReport.BATTALION_DICT["EMS"])
        detailed = self.detailed_personnel()
        return {"detailed": detailed}

    
class FullRosterReport:
    def __init__(self, *strategies: ReportType):
        self.strategies = strategies
        
    def get_report(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        report_data = []
        
        for strategy in self.strategies:
            strategy_data = strategy.generate(data)
            report_data.extend(strategy_data)
            
        return report_data


class FormulaIDAudit(ReportType):
    def verify_specialties(self):
        employees = {
            "CORRECT": dict(),
            "INCORRECT": dict()
        }
        for record in self.data.values():
            is_field = self.is_field_compliment(record)

            if is_field:
                company = self.get_specialty_company(record)
                shift = self._get_shift_specialty(record)
                first_name, last_name = record["first_name"], record['last_name']

                if company != 'None' and shift != 'None':
                    employees["CORRECT"].update({last_name+', '+first_name: (company,shift)})
                    ...
                else:
                    employees["INCORRECT"].update({last_name+', '+first_name: (company,shift)})

            else:
                company = self.get_specialty_company(record)
                shift = self._get_shift_specialty(record)
                first_name, last_name = record["first_name"], record['last_name']

                if company == 'None' and  shift in ['None', None]:
                    ...
                else:
                    employees["INCORRECT"].update({last_name+', '+first_name: (company,shift)})

        return employees

    
    def __call__(self) -> Any:
        employees = self.verify_specialties()
        return {'employees': employees}
        
class RosterAudit(ReportType):
    def check_for_duplicate_name_on_roster(self):
        CURRENT_SHIFT = self.get_current_shift(self.data)
        name_list = list()
        duplicate_list = dict()
        for record in self.data.values():
            is_field = self.is_field_compliment(record)
            is_right_now = self.right_now(record)

            speciality_shift = self._get_shift_specialty(record)
            
            if is_field and is_right_now and speciality_shift == CURRENT_SHIFT:
                name = record["name"]
                if name not in name_list:
                    name_list.append(name)
                else:
                    duplicate_list[name] = ({"paycode":record["paycode"], "company": record["company"], "region": record["shift_start"], 'rank': record["shift_end"]})
        return duplicate_list
    def __call__(self, *args: Any, **kwds: Any) -> Any:
        duplicates = self.check_for_duplicate_name_on_roster()
        return {"duplicated": duplicates}
    

class OvertimeAudit(ReportType):
    def get_current_overtime(self):
        overtime = dict()
        for record in self.data.values():
            is_field = self.region_in_field(record)
            is_right_now = self.right_now(record)
            paycode = self.get_paycode(record)
            start, end = self.ISO_8601_reformatter((record["shift_start"], record["shift_end"]))

            if is_field and is_right_now:
                if re.search(r"OTCB|OTHO|OTOR|EXT", paycode):
                    overtime[record['name']] = {"company": record["company"],"paycode": paycode,"shift_start": start, "shift_end": end}
        return overtime

    def __call__(self, *args: Any, **kwds: Any) -> Any:
        overtime = self.get_current_overtime()
        return {"overtime": overtime}

    
    ...


class PositionReport(ReportType):

    def get_positions(self):
        company_positions = ["Fire Lieutenant", "Fire Driver", "Fire Private", "Fire Private"]
        positions = dict()
        for record in self.data.values():
            company = record["company"]
            position = record["rank"]
            
            if company not in positions.keys():
                positions[company] = []
            positions[company] = company_positions 

        print(positions)
        return positions
    def __call__(self, *args: Any, **kwds: Any) -> Any:
        data = self.get_positions()
        return {'data': data}


